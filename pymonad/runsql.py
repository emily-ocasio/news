"""
Intent and eliminator for SQL (database) effects.
"""

# pylint:disable=W0212
from dataclasses import dataclass
from typing import TypeVar
from typing import Any, cast

import pandas as pd
import duckdb
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule
from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter

from .array import Array
from .monad import Unit
from .environment import DbBackend, Environment
from .run import Run, _unhandled, ask, local, throw, ErrorPayload
from .string import String


class SQL(String):
    """
    Represents a SQL string.
    """


type SQLParam = String | int | float | None


class SQLParams(Array[SQLParam]):
    """
    Represents a list of SQL parameters.
    """

    def to_params(self) -> tuple:
        """
        Convert SQLParams to a tuple of parameters.
        """

        def _convert(param: SQLParam) -> str | int | float | None:
            match param:
                case String(s):
                    return s
                case _:
                    return param

        return tuple(_convert(param) for param in self.a)


class SQLExecutionError(Exception):
    """
    Raised by run_sql performer when the underlying
    driver raises an exception. Captures SQL and params.
    """

    def __init__(self, sql: str, params: tuple | None, original: Exception):
        self.sql = sql
        self.params = params
        self.original = original
        super().__init__(str(original))

    def __str__(self) -> str:
        p = "None" if self.params is None else repr(self.params)
        return (
            f"{self.original.__class__.__name__}: {self.original}\n"
            f"--- SQL Context --------------------------------\n"
            f"SQL:\n{self.sql}\n"
            f"Params: {p}\n"
            f"------------------------------------------------"
        )


# --- DB intents ---
@dataclass(frozen=True)
class SqlQuery:
    """
    Represents a SQL query with parameters.
    """

    sql: SQL
    params: SQLParams = SQLParams(())


@dataclass(frozen=True)
class SqlExec:
    """
    Represents a SQL execution command with parameters.
    """

    sql: SQL
    params: SQLParams = SQLParams(())


@dataclass(frozen=True)
class SqlScript:
    """
    Represents a SQL script with multiple statements.
    """

    sql: SQL  # multi-statement DDL/DML


@dataclass(frozen=True)
class InTransaction:
    """
    Represents a database transaction.
    """

    program: Run[Any]  # sub-program run atomically


@dataclass(frozen=True)
class SqlExport:
    """
    Intent to export results of a SQL query into a spreadsheet (or fallback CSV),
    optionally with row-band formatting by group.
    """

    sql: SQL
    filename: str
    sheet: str | None = None
    band_by_group_col: str | None = None
    band_wrap: int = 2  # how many alternating bands/colors


@dataclass(frozen=True)
class SqlImport:
    """
    Intent to import a spreadsheet (or CSV) into DuckDB.
    """

    filename: str
    table_name: str
    sheet: str | None = None


@dataclass(frozen=True)
class SqlRegister:
    """
    Intent to register a pandas DataFrame with DuckDB.
    """
    name: str
    df: pd.DataFrame


def sql_query(sql: SQL, params: SQLParams = SQLParams(())) -> Run[Array]:
    """
    Smart constructor for SQL queries with intent.
    """
    return Run(lambda self: self._perform(SqlQuery(sql, params), self), _unhandled)


def sql_exec(sql: SQL, params: SQLParams = SQLParams(())) -> Run[Unit]:
    """
    Smart constructor for SQL execution with intent.
    """
    return Run(lambda self: self._perform(SqlExec(sql, params), self), _unhandled)


def sql_script(sql: SQL) -> "Run[None]":
    """
    Smart constructor for SQL scripts with intent.
    """
    return Run(lambda self: self._perform(SqlScript(sql), self), _unhandled)


def in_transaction(prog: "Run[A]") -> "Run[A]":
    """
    Smart constructor for database transactions with intent.
    """
    return Run(lambda self: self._perform(InTransaction(prog), self), _unhandled)


def sql_export(
    sql: SQL,
    filename: str,
    sheet: str | None = None,
    band_by_group_col: str | None = None,
    band_wrap: int = 2,
) -> Run[None]:
    """
    Smart constructor for exporting SQL query results to a spreadsheet (or CSV).
    """
    return Run(
        lambda self: self._perform(
            SqlExport(sql, filename, sheet, band_by_group_col, band_wrap),
            self,
        ),
        _unhandled,
    )


def sql_import(
    filename: str,
    table_name: str,
    sheet: str | None = None,
) -> Run[None]:
    """
    Smart constructor for importing a spreadsheet (or CSV) into DuckDB.
    """
    return Run(
        lambda self: self._perform(SqlImport(filename, table_name, sheet), self),
        _unhandled,
    )


def sql_register(name: str, df: pd.DataFrame) -> Run[None]:
    """
    Smart constructor for registering a pandas DataFrame in DuckDB.
    """
    return Run(lambda self: self._perform(SqlRegister(name, df), self), _unhandled)


A = TypeVar("A")

def _sql_export(df: pd.DataFrame, filename, sheet, band_by_group_col, band_wrap):
    widths = _compute_column_widths(df)
    try:
        with pd.ExcelWriter(filename, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name=sheet or "Sheet1", index=False)
        # Apply formatting immediately after successful Excel write
        if band_by_group_col:
            _apply_band_formatting_xlsx(
                filename,
                sheet or "Sheet1",
                band_by_group_col,
                band_wrap,
            )
            _auto_fit_columns_xlsx(filename, sheet or "Sheet1", widths)
        else:
            _set_freeze_panes_xlsx(filename, sheet or "Sheet1")
            _auto_fit_columns_xlsx(filename, sheet or "Sheet1", widths)
    except (
        PermissionError,
        FileNotFoundError,
        OSError,
        ValueError,
    ):
        # Fallback to CSV if Excel write fails
        fallback = (
            filename
            if filename.lower().endswith(".csv")
            else (filename + ".csv")
        )
        df.to_csv(fallback, index=False)


def _sql_import(filename: str, sheet: str | None) -> pd.DataFrame:
    if filename.lower().endswith((".xlsx", ".xlsm", ".xls")):
        if sheet is None:
            return pd.read_excel(filename, sheet_name=0)
        return pd.read_excel(filename, sheet_name=sheet)
    if filename.lower().endswith(".csv"):
        return pd.read_csv(filename)
    raise ValueError(f"Unsupported import file type: {filename}")


def _apply_band_formatting_xlsx(
    filename: str, sheet: str, group_col: str, band_wrap: int = 2
) -> None:
    """
    After writing the Excel file, color rows in bands.
    Generic behavior:
      - If `group_col` already contains integer band indices in 0..band_wrap-1,
        apply conditional formats directly from that column (no helper col).
      - Otherwise, insert a helper column that alternates a band index whenever
        the `group_col` value changes (classic grouping bands).
    """
    try:
        wb = load_workbook(filename)
    except KeyError as e:
        print(f"[WARN] Cannot load {filename} for band formatting: {e}")
        return
    if sheet not in wb.sheetnames:
        wb.close()
        return
    ws = wb[sheet]

    # Read header row
    header = [cell.value for cell in ws[1]]
    if group_col not in header:
        wb.close()
        return

    # locate the (1-based) column index for the requested group_col
    grp_idx = header.index(group_col) + 1

    # Try DIRECT MODE: if the band column already contains explicit integers
    # within [0, band_wrap), use them as-is and apply per-index formulas.
    try:
        grp_col_letter_direct = cast(Cell, ws.cell(row=1, column=grp_idx)).column_letter
        sample_rows = min(ws.max_row, 101)
        seen_any = False
        direct_ok = True
        for row in range(2, sample_rows + 1):
            val = ws.cell(row=row, column=grp_idx).value
            if val is None or val == "":
                continue
            seen_any = True
            try:
                ival = int(str(val))
                if ival < 0 or ival >= band_wrap:
                    direct_ok = False
                    break
            except Exception:   # pylint: disable=W0718
                direct_ok = False
                break
        if seen_any and direct_ok:
            if ws.max_row < 2:
                ws.freeze_panes = 'A2'
                wb.save(filename)
                wb.close()
                return
            last_col = ws.max_column
            last_col_letter = cast(Cell, ws.cell(row=1, column=last_col)).column_letter
            data_range = f"A2:{last_col_letter}{ws.max_row}"
            # simple 3-color palette; cycles up to band_wrap
            fill_list = [
                PatternFill(
                    start_color="FFEEEE", end_color="FFEEEE", fill_type="solid"
                ),
                PatternFill(
                    start_color="EEFFEE", end_color="EEFFEE", fill_type="solid"
                ),
                PatternFill(
                    start_color="EEEEFF", end_color="EEEEFF", fill_type="solid"
                ),
            ]
            max_wrap = min(band_wrap, len(fill_list))
            for i in range(max_wrap):
                formula = f"${grp_col_letter_direct}2 = {i}"
                ws.conditional_formatting.add(
                    data_range,
                    FormulaRule(formula=[formula], fill=fill_list[i]),
                )
            if group_col.startswith("__"):
                ws.column_dimensions[grp_col_letter_direct].hidden = True
            ws.freeze_panes = 'A2'
            wb.save(filename)
            wb.close()
            return
    except Exception:  # pylint: disable=W0718
        # fall through to helper-based mode
        pass

    # HELPER MODE: Insert helper column A with a rolling band counter that
    # increments when group_col value changes (alternating bands by group).
    ws.insert_cols(1)
    ws.cell(row=1, column=1, value="_band_helper")

    # After insert, group_col shifts right by one
    grp_col_letter = cast(Cell, ws.cell(row=1, column=grp_idx + 1)).column_letter
    helper_letter = cast(Cell, ws.cell(row=1, column=1)).column_letter

    # Fill helper formula
    for row in range(2, ws.max_row + 1):
        cell = cast(Cell, ws.cell(row=row, column=1))
        formula = (
            (
                f"=MOD(IF(${grp_col_letter}{row} = ${grp_col_letter}{row-1}, "
                f"${helper_letter}{row-1}, ${helper_letter}{row-1} + 1), {band_wrap})"
            )
            if row > 2
            else "=0"
        )
        cell.value = formula

    # Pattern fills
    fill_list = [
        PatternFill(start_color="FFEEEE", end_color="FFEEEE", fill_type="solid"),
        PatternFill(start_color="EEFFEE", end_color="EEFFEE", fill_type="solid"),
        PatternFill(start_color="EEEEFF", end_color="EEEEFF", fill_type="solid"),
    ]
    max_wrap = min(band_wrap, len(fill_list))

    # Apply conditional formatting across all columns A -> last column
    last_col = ws.max_column
    last_col_letter = cast(Cell, ws.cell(row=1, column=last_col)).column_letter
    if ws.max_row < 2:
        if group_col.startswith("__"):
            ws.column_dimensions[grp_col_letter].hidden = True
        ws.freeze_panes = 'A2'
        wb.save(filename)
        wb.close()
        return
    for i in range(max_wrap):
        formula = f"${helper_letter}2 = {i}"
        ws.conditional_formatting.add(
            f"A2:{last_col_letter}{ws.max_row}",
            FormulaRule(formula=[formula], fill=fill_list[i]),
        )

    if group_col.startswith("__"):
        ws.column_dimensions[grp_col_letter].hidden = True
    ws.freeze_panes = 'A2'
    wb.save(filename)
    wb.close()


def _estimate_text_width(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, float) and pd.isna(value):
        return 0.0
    text = str(value)
    if not text:
        return 0.0
    max_line = max(text.splitlines(), key=len)
    width = 0.0
    for ch in max_line:
        if "A" <= ch <= "Z":
            width += 0.9
        elif "a" <= ch <= "z":
            width += 0.8
        elif "0" <= ch <= "9":
            width += 1.0
        elif ch.isspace():
            width += 0.5
        else:
            width += 1.0
    return width


def _compute_column_widths(
    df: pd.DataFrame,
    sample_limit: int = 1000,
    padding: float = 2.0,
    max_width: float = 60.0,
    min_width: float = 5.0,
) -> list[float]:
    widths: list[float] = []
    for col in df.columns:
        series = df[col]
        if len(series) > sample_limit:
            half = max(sample_limit // 2, 1)
            sample = pd.concat([series.head(half), series.tail(half)])
        else:
            sample = series
        estimates: list[float] = []
        for val in sample:
            est = _estimate_text_width(val)
            if est > 0:
                estimates.append(est)
        if not estimates:
            widths.append(0.0)
            continue
        estimates.sort()
        cutoff_idx = int(0.97 * (len(estimates) - 1))
        base = estimates[cutoff_idx]
        width = min(base + padding, max_width)
        widths.append(max(width, min_width))
    return widths


def _auto_fit_columns_xlsx(filename: str, sheet: str, widths: list[float]) -> None:
    try:
        wb = load_workbook(filename)
    except KeyError as e:
        print(f"[WARN] Cannot load {filename} for auto-fit: {e}")
        return
    if sheet not in wb.sheetnames:
        wb.close()
        return
    ws = wb[sheet]
    has_helper = ws.cell(row=1, column=1).value == "_band_helper"
    if has_helper:
        ws.column_dimensions["A"].width = 4
    start_col = 2 if has_helper else 1
    for idx, width in enumerate(widths, start=start_col):
        ws.column_dimensions[get_column_letter(idx)].width = float(width)
    wb.save(filename)
    wb.close()


def _set_freeze_panes_xlsx(filename: str, sheet: str) -> None:
    try:
        wb = load_workbook(filename)
        if sheet in wb.sheetnames:
            ws = wb[sheet]
            ws.freeze_panes = 'A2'  # Freeze the first row
            wb.save(filename)
        wb.close()
    except Exception as e:
        print(f"[WARN] Could not set freeze panes: {e}")
        raise e


def _get_backend_and_conn_run() -> Run[tuple[DbBackend, Any]]:
    def resolve(env: Environment) -> Run[tuple[DbBackend, Any]]:
        backend = env["current_backend"]
        con = env["connections"].get(backend)
        if con is None:
            return throw(ErrorPayload(f"No connection for backend: {backend}"))
        return Run.pure((backend, con))
    return ask() >> resolve


def _duckdb_query_dicts(
    con: duckdb.DuckDBPyConnection, sql_s: str, params: SQLParams
) -> Array[dict]:
    try:
        cur = con.execute(sql_s, params.to_params())
        cols = [d[0] for d in (cur.description or [])]
        rows = cur.fetchall()
        return Array(tuple(dict(zip(cols, tup)) for tup in rows))
    except Exception as ex:  # noqa: BLE001
        raise SQLExecutionError(sql_s, params.to_params(), ex) from ex


def run_sql(prog: Run[A]) -> Run[A]:
    """
    Eliminator for SQL database calls. Routes to the current backend from env.
    """

    def step(self_run: Run[Any]) -> A:
        parent = self_run._perform

        def perform(intent: Any, current: "Run[Any]") -> Any:
            match intent:
                case SqlQuery(sql, params):
                    backend, con = _get_backend_and_conn_run()._step(current)
                    if backend == DbBackend.SQLITE:
                        try:
                            cur = con.execute(str(sql), params.to_params())
                            rows = cur.fetchall()
                            cur.close()
                            return rows
                        except Exception as ex:  # noqa: BLE001
                            raise SQLExecutionError(
                                str(sql), params.to_params(), ex
                            ) from ex
                    return _duckdb_query_dicts(con, str(sql), params)

                case SqlExec(sql, params):
                    backend, con = _get_backend_and_conn_run()._step(current)
                    try:
                        con.execute(str(sql), params.to_params())
                        if backend == DbBackend.SQLITE:
                            con.commit()
                        return None
                    except Exception as ex:  # noqa: BLE001
                        raise SQLExecutionError(
                            str(sql), params.to_params(), ex
                        ) from ex

                case SqlScript(sql):
                    backend, con = _get_backend_and_conn_run()._step(current)
                    try:
                        if backend == DbBackend.SQLITE:
                            con.executescript(str(sql))
                        else:
                            con.execute(str(sql))
                        return None
                    except Exception as ex:  # noqa: BLE001
                        raise SQLExecutionError(str(sql), None, ex) from ex

                case SqlExport(sql, filename, sheet, band_by_group_col, band_wrap):
                    backend, con = _get_backend_and_conn_run()._step(current)
                    try:
                        if backend == DbBackend.SQLITE:
                            df = pd.read_sql_query(str(sql), con)
                        else:
                            df = con.execute(str(sql)).df()
                    except Exception as ex:
                        raise SQLExecutionError(str(sql), None, ex) from ex
                    _sql_export(df, filename, sheet, band_by_group_col, band_wrap)
                    return None
                case SqlImport(filename, table_name, sheet):
                    backend, con = _get_backend_and_conn_run()._step(current)
                    if backend != DbBackend.DUCKDB:
                        raise SQLExecutionError(
                            f"IMPORT {filename}",
                            None,
                            RuntimeError("SqlImport requires DuckDB backend."),
                        )
                    try:
                        df = _sql_import(filename, sheet)
                        con.register(table_name, df)
                        cols = [c for c in df.columns if c != "_band_helper"]
                        if cols:
                            select_cols = ", ".join(f'"{c}"' for c in cols)
                        else:
                            select_cols = "*"
                        con.execute(
                            f'CREATE OR REPLACE TABLE "{table_name}" AS SELECT'
                            f' {select_cols} FROM "{table_name}";'
                        )
                        return None
                    except Exception as ex:  # noqa: BLE001
                        raise SQLExecutionError(
                            f"IMPORT {filename}", None, ex
                        ) from ex
                case SqlRegister(name, df):
                    backend, con = _get_backend_and_conn_run()._step(current)
                    if backend != DbBackend.DUCKDB:
                        raise SQLExecutionError(
                            f"REGISTER {name}",
                            None,
                            RuntimeError("SqlRegister requires DuckDB backend."),
                        )
                    try:
                        con.register(name, df)
                        return None
                    except Exception as ex:  # noqa: BLE001
                        raise SQLExecutionError(f"REGISTER {name}", None, ex) from ex

                case InTransaction(subprog):
                    backend, con = _get_backend_and_conn_run()._step(current)
                    try:
                        if backend == DbBackend.DUCKDB:
                            con.execute("BEGIN")
                        result = subprog._step(current)
                        if backend == DbBackend.DUCKDB:
                            con.execute("COMMIT")
                        else:
                            con.commit()
                        return result
                    except Exception:
                        if backend == DbBackend.DUCKDB:
                            con.execute("ROLLBACK")
                        else:
                            con.rollback()
                        raise

                case _:
                    return parent(intent, current)

        inner = Run(prog._step, perform)
        return inner._step(inner)

    return Run(step, lambda i, c: c._perform(i, c))


def with_duckdb(subprog: Run[A]) -> Run[A]:
    """
    Context wrapper to run a sub-program in DuckDB mode.
    """
    return local(
        lambda env: {**env, "current_backend": DbBackend.DUCKDB},
        subprog,
    )
